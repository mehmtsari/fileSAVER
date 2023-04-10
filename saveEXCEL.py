


class saveEXCEL:
    def __init__(self,FileNAME:str, WorkBookTitle : str,justread:bool):
        """filename : Must be 'filename.xlsx'
           WorkBookTitle : WorkSheet Name
           justread : İf True; You just Read / İf False; You just Write
        """
        from openpyxl import Workbook,load_workbook
        self.FileNAME = FileNAME
        if self.FileNAME.endswith('.xlsx'):
            self.FileNAME += '.xlsx'
            
        if not justread:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = WorkBookTitle
        else:
            self.wb = load_workbook(self.FileNAME)
            self.ws =  self.wb[WorkBookTitle]
        
    def add_row(self,columns:list):
        self.ws.append(columns)
        
    def read_rows(self):
        """Return Rows [list]
        """
        max_row = self.ws.max_row
        rows = []
        for row_index in range(1,max_row+1):
            if all([cell.value is None for cell in self.ws[row_index]]):
                continue
            row_values = []
            for cell in self.ws[row_index]:
                row_values.append(cell.value)
            rows.append(row_values)
        return rows
    
    def saveEXCEL(self):
        self.wb.save(filename=self.FileNAME)
        


