
import csv
class saveCSV:
    def __init__(self,filepath:str,justreading:bool = False) -> None:
        
        self.filePATH = filepath
        if self.filePATH.endswith('.csv'):
            self.filePATH += '.csv'
        if not justreading:
            self.file = open(filepath,mode='w',newline='')
            self.writer = csv.writer(self.file,delimiter=',',quotechar='"',quoting=csv.QUOTE_MINIMAL)
        else:
            self.file = open(filepath,mode='r')
            self.reader = csv.reader(self.file,delimiter=',',quotechar='"',quoting=csv.QUOTE_MINIMAL)
            
    def add_row(self,row):
        self.writer.writerow(row)
    
    def add_rows(self,rows):
        self.writer.writerows(rows)
    
    def read_row(self,rowcount):
        rows = [row for row in self.reader]
        rowcount -= 1 
        if rowcount > len(rows):
            print('Satır üzerinde herhangi bir bilgi bulunmamakta.')
            return None
        return rows[rowcount]
            
        
    def read_rows(self):
        return [row for row in self.reader]     
    
    def close(self):
        self.file.close()


from openpyxl import Workbook,load_workbook
class saveEXCEL:
    def __init__(self,FileNAME:str, WorkBookTitle : str,justread:bool):
        """filename : Must be 'filename.xlsx'
           WorkBookTitle : WorkSheet Name
           justread : İf True; You just Read / İf False; You just Write
        """
        
        self.FileNAME = FileNAME
        if self.FileNAME.endswith('.xlsx'):
            self.FileNAME += '.xlsx'
            
        if not justread:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = WorkBookTitle
        else:
            self.wb = load_workbook(self.FileNAME)
            self.ws =  self.wb[WorkBookTitle]
        
    def add_row(self,columns:list):
        self.ws.append(columns)
        
    def read_rows(self):
        """Return Rows [list]
        """
        max_row = self.ws.max_row
        rows = []
        for row_index in range(1,max_row+1):
            if all([cell.value is None for cell in self.ws[row_index]]):
                continue
            row_values = []
            for cell in self.ws[row_index]:
                row_values.append(cell.value)
            rows.append(row_values)
        return rows
    
    def saveEXCEL(self):
        self.wb.save(filename=self.FileNAME)
        

class saveTXT:
    def __init__(self,filename:str,mode:str) -> None:
        """filename : Must be 'filename.txt'
        mode ; 
            'w' : Open new one
            'a' : if have, read. dont have, create
            'r' : just read
        """
        self.file_name = filename
        self.file = open(filename,mode)
        
    def add_row(self,Text):
        self.file.write(Text)
        self.file.write('\n')
        
    def add_rows(self, RowList:list):
        for item in RowList:
            self.file.write(item)
            self.file.write('\n')
    
    def read_all(self):
        return self.file.read()
    
    def read_lines(self):
        return self.file.readlines()

    def saveTXT(self):
        """Closing the TXT file
        """
        self.file.close()


import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
class saveXML:
    def __init__(self,filePATH,justreading:bool,tree_name:str = None) -> None:
        
        self.mode = justreading
        self.file_path = filePATH
        if not justreading:
            self.XMLTree = ET.Element(tree_name)
        
    def read(self):
        with open(self.file_path, 'r', encoding='utf-8') as f:
            return BeautifulSoup(f.read(),'xml')
        
    def find_all(self,element_name):
        with open(self.file_path, 'r', encoding='utf-8') as f:
            return BeautifulSoup(f.read(),'xml').find_all(element_name)
    
    def create_element(self,element_name: str, elements: list = None, just_return = False):
        """
        Creates an XML element with the given element_name and adds its sub-elements (elements).
        If no sub-elements are specified, only the main element is created.

        :param element_name: The name of the XML element to create. Must be a string.
        :param elements: The sub-elements of the element. If desired, can be specified as a tuple in the form (sub_element_name, [sub_element_elements]).
        Each sub_element_elements list creates a new sub-element with the name sub_element_name that will be located under the sub-elements.
        The elements list is a list of these tuples and can be used to create multiple sub-elements at the same time.
        For example: [('element_1', 'element_name'), ('element_2', [('sub_element_1', []), ('sub_element_2',[(down_element_1,'element_name')])])]
        :param just_return: If set to True, the element is created but not stored, and only the created element is returned.
        By default, its value is False, which means the created element is stored and an element object is returned.
        :return: The created XML element.
        """
        if self.mode:
            raise Exception('just_reading mode Active')

        element = ET.Element(element_name)
        
        if elements:
            for item in elements:
                if isinstance(item[0], str) and isinstance(item[1],list):
                    sub_element_name = item[0]
                    sub_elements = item[1]
                    sub_element = self.create_element(sub_element_name, sub_elements,True)
                    element.append(sub_element)
                else:
                    tag_name = item[0]
                    tag_text = item[1]
                    ET.SubElement(element, tag_name).text = tag_text
                    
        if just_return:
            return element
        else:
            self.XMLTree.append(element)
            return element

    def saveXML(self):
        if self.mode:
            raise Exception('just_reading mode Active')
        with open(self.file_path, 'wb') as f:
            f.write(ET.tostring(self.XMLTree, encoding='utf-8', method='xml'))

